import io
import os
from datetime import datetime, timezone
from functools import wraps

from bson import ObjectId
from flask import Flask, jsonify, redirect, render_template, request, session, url_for, send_file
from pymongo import MongoClient, ASCENDING
from werkzeug.security import check_password_hash, generate_password_hash
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-key")

# MongoDB connection pool
client = MongoClient(os.environ.get("MONGO_URI", "mongodb://localhost:27017/worklogger"))
db = client.get_default_database() if "mongodb+srv" in os.environ.get("MONGO_URI", "") else client["worklogger"]

tasks_col = db["tasks"]
learnings_col = db["learnings"]
users_col = db["users"]
holidays_col = db["holidays"]

# Ensure indexes
tasks_col.create_index([("date", ASCENDING)])
learnings_col.create_index([("date", ASCENDING)])
users_col.create_index("username", unique=True)
holidays_col.create_index("date", unique=True)

# ─── Seed default user ────────────────────────────────────────────────────────

def seed_user():
    if not users_col.find_one({"username": "yogendr023"}):
        users_col.insert_one({
            "username": "yogendr023",
            "password": generate_password_hash("yogendr023password"),
            "created_at": datetime.now(timezone.utc),
        })

seed_user()


# ─── Auth helpers ─────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def serialize_doc(doc):
    doc["_id"] = str(doc["_id"])
    for key, val in doc.items():
        if isinstance(val, datetime):
            doc[key] = val.isoformat()
    return doc


# ─── Auth routes ──────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user"):
        return redirect(url_for("index"))
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = users_col.find_one({"username": username})
        if user and check_password_hash(user["password"], password):
            session["user"] = username
            return redirect(url_for("index"))
        error = "Invalid username or password."
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/api/change-password", methods=["POST"])
@login_required
def change_password():
    data = request.get_json()
    if not data or not data.get("current_password") or not data.get("new_password"):
        return jsonify({"error": "All fields are required."}), 400
    user = users_col.find_one({"username": session["user"]})
    if not user or not check_password_hash(user["password"], data["current_password"]):
        return jsonify({"error": "Current password is incorrect."}), 403
    if len(data["new_password"]) < 6:
        return jsonify({"error": "New password must be at least 6 characters."}), 400
    users_col.update_one(
        {"username": session["user"]},
        {"$set": {"password": generate_password_hash(data["new_password"])}},
    )
    return jsonify({"ok": True})


# ─── Pages ────────────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def index():
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return render_template("index.html", today=today, user=session["user"])


# ─── Tasks API ────────────────────────────────────────────────────────────────

@app.route("/api/tasks", methods=["GET"])
@login_required
def get_tasks():
    date = request.args.get("date")
    if not date:
        return jsonify({"error": "date parameter required"}), 400
    docs = list(tasks_col.find({"date": date}, sort=[("created_at", ASCENDING)]))
    return jsonify([serialize_doc(d) for d in docs])


@app.route("/api/tasks", methods=["POST"])
@login_required
def create_task():
    data = request.get_json()
    if not data or not data.get("date") or not data.get("title"):
        return jsonify({"error": "date and title required"}), 400
    now = datetime.now(timezone.utc)
    doc = {
        "date": data["date"],
        "title": data["title"].strip(),
        "note": data.get("note", "").strip(),
        "status": data.get("status", "done"),
        "created_at": now,
        "updated_at": now,
    }
    result = tasks_col.insert_one(doc)
    doc["_id"] = result.inserted_id
    return jsonify(serialize_doc(doc)), 201


@app.route("/api/tasks/<task_id>", methods=["PUT"])
@login_required
def update_task(task_id):
    data = request.get_json()
    if not data:
        return jsonify({"error": "no data provided"}), 400
    updates = {"updated_at": datetime.now(timezone.utc)}
    if "title" in data:
        updates["title"] = data["title"].strip()
    if "note" in data:
        updates["note"] = data["note"].strip()
    if "status" in data and data["status"] in ("done", "in_progress"):
        updates["status"] = data["status"]
    result = tasks_col.find_one_and_update(
        {"_id": ObjectId(task_id)},
        {"$set": updates},
        return_document=True,
    )
    if not result:
        return jsonify({"error": "not found"}), 404
    return jsonify(serialize_doc(result))


@app.route("/api/tasks/<task_id>", methods=["DELETE"])
@login_required
def delete_task(task_id):
    result = tasks_col.delete_one({"_id": ObjectId(task_id)})
    if result.deleted_count == 0:
        return jsonify({"error": "not found"}), 404
    return jsonify({"deleted": task_id})


@app.route("/api/tasks/<task_id>/carry-forward", methods=["POST"])
@login_required
def carry_forward_task(task_id):
    data = request.get_json()
    target_date = (data or {}).get("target_date", "").strip()
    if not target_date:
        return jsonify({"error": "target_date required"}), 400
    try:
        datetime.strptime(target_date, "%Y-%m-%d")
    except ValueError:
        return jsonify({"error": "invalid date format"}), 400
    original = tasks_col.find_one({"_id": ObjectId(task_id)})
    if not original:
        return jsonify({"error": "task not found"}), 404
    now = datetime.now(timezone.utc)
    new_doc = {
        "date":         target_date,
        "title":        original["title"],
        "note":         original.get("note", ""),
        "status":       "in_progress",
        "carried_from": str(original["_id"]),
        "created_at":   now,
        "updated_at":   now,
    }
    result = tasks_col.insert_one(new_doc)
    new_doc["_id"] = result.inserted_id
    return jsonify(serialize_doc(new_doc)), 201


# ─── Learnings API ────────────────────────────────────────────────────────────

@app.route("/api/learnings", methods=["GET"])
@login_required
def get_learnings():
    date = request.args.get("date")
    if not date:
        return jsonify({"error": "date parameter required"}), 400
    docs = list(learnings_col.find({"date": date}, sort=[("created_at", ASCENDING)]))
    return jsonify([serialize_doc(d) for d in docs])


@app.route("/api/learnings", methods=["POST"])
@login_required
def create_learning():
    data = request.get_json()
    if not data or not data.get("date") or not data.get("content"):
        return jsonify({"error": "date and content required"}), 400
    now = datetime.now(timezone.utc)
    tags = [t.strip() for t in data.get("tags", []) if t.strip()]
    doc = {
        "date": data["date"],
        "content": data["content"].strip(),
        "tags": tags,
        "created_at": now,
        "updated_at": now,
    }
    result = learnings_col.insert_one(doc)
    doc["_id"] = result.inserted_id
    return jsonify(serialize_doc(doc)), 201


@app.route("/api/learnings/<learning_id>", methods=["PUT"])
@login_required
def update_learning(learning_id):
    data = request.get_json()
    if not data:
        return jsonify({"error": "no data provided"}), 400
    updates = {"updated_at": datetime.now(timezone.utc)}
    if "content" in data:
        updates["content"] = data["content"].strip()
    if "tags" in data:
        updates["tags"] = [t.strip() for t in data["tags"] if t.strip()]
    result = learnings_col.find_one_and_update(
        {"_id": ObjectId(learning_id)},
        {"$set": updates},
        return_document=True,
    )
    if not result:
        return jsonify({"error": "not found"}), 404
    return jsonify(serialize_doc(result))


@app.route("/api/learnings/<learning_id>", methods=["DELETE"])
@login_required
def delete_learning(learning_id):
    result = learnings_col.delete_one({"_id": ObjectId(learning_id)})
    if result.deleted_count == 0:
        return jsonify({"error": "not found"}), 404
    return jsonify({"deleted": learning_id})


@app.route("/review")
@login_required
def review():
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return render_template("review.html", today=today)


@app.route("/stats")
@login_required
def stats_page():
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return render_template("stats.html", today=today)


@app.route("/api/stats")
@login_required
def api_stats():
    import datetime as dt_module
    from calendar import monthrange

    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    try:
        year  = int(request.args.get("year",  today_str[:4]))
        month = int(request.args.get("month", today_str[5:7]))
    except (ValueError, TypeError):
        return jsonify({"error": "invalid year/month"}), 400

    first_day    = f"{year:04d}-{month:02d}-01"
    last_day_num = monthrange(year, month)[1]
    last_day     = f"{year:04d}-{month:02d}-{last_day_num:02d}"

    month_tasks     = tasks_col.count_documents({"date": {"$gte": first_day, "$lte": last_day}})
    month_learnings = learnings_col.count_documents({"date": {"$gte": first_day, "$lte": last_day}})

    task_dates     = {d["date"] for d in tasks_col.find({}, {"date": 1})}
    learning_dates = {d["date"] for d in learnings_col.find({}, {"date": 1})}
    all_logged     = sorted(task_dates | learning_dates)

    holidays_set = {h["date"] for h in holidays_col.find({})}

    today_date = dt_module.date.fromisoformat(today_str)
    logged_set = set(all_logged)

    def current_streak():
        streak = 0
        cur = today_date
        for _ in range(1826):  # max 5 years back
            day_str    = cur.isoformat()
            is_weekend = cur.weekday() >= 5
            is_holiday = day_str in holidays_set
            if day_str in logged_set:
                streak += 1
            elif is_weekend or is_holiday:
                pass
            else:
                break
            cur -= dt_module.timedelta(days=1)
        return streak

    def longest_streak():
        if not all_logged:
            return 0
        start = dt_module.date.fromisoformat(all_logged[0])
        best = run = 0
        cur = start
        while cur <= today_date:
            day_str    = cur.isoformat()
            is_weekend = cur.weekday() >= 5
            is_holiday = day_str in holidays_set
            if day_str in logged_set:
                run += 1
                best = max(best, run)
            elif is_weekend or is_holiday:
                pass
            else:
                run = 0
            cur += dt_module.timedelta(days=1)
        return best

    month_prefix    = f"{year:04d}-{month:02d}-"
    month_logged    = sorted(d for d in all_logged   if d.startswith(month_prefix))
    month_holidays  = sorted(h for h in holidays_set if h.startswith(month_prefix))

    return jsonify({
        "days_logged":    month_logged,
        "holidays":       month_holidays,
        "streak":         current_streak(),
        "longest_streak": longest_streak(),
        "month_tasks":    month_tasks,
        "month_learnings": month_learnings,
        "year":  year,
        "month": month,
    })


@app.route("/api/holidays/toggle", methods=["POST"])
@login_required
def toggle_holiday():
    data = request.get_json()
    date_str = (data or {}).get("date", "").strip()
    if not date_str:
        return jsonify({"error": "date required"}), 400
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return jsonify({"error": "invalid date format"}), 400
    if holidays_col.find_one({"date": date_str}):
        holidays_col.delete_one({"date": date_str})
        return jsonify({"action": "removed", "date": date_str})
    holidays_col.insert_one({"date": date_str})
    return jsonify({"action": "added", "date": date_str})


@app.route("/api/review")
@login_required
def api_review():
    date_from = request.args.get("from")
    date_to   = request.args.get("to")
    if not date_from or not date_to:
        return jsonify({"error": "from and to required"}), 400
    if date_from > date_to:
        return jsonify({"error": "from must be <= to"}), 400

    q = request.args.get("q", "").strip()
    try:
        page = max(1, int(request.args.get("page", 1)))
    except (ValueError, TypeError):
        page = 1

    PAGE_SIZE = 7

    task_query = {"date": {"$gte": date_from, "$lte": date_to}}
    learning_query = {"date": {"$gte": date_from, "$lte": date_to}}
    if q:
        task_query["$or"] = [
            {"title": {"$regex": q, "$options": "i"}},
            {"note":  {"$regex": q, "$options": "i"}},
        ]
        learning_query["content"] = {"$regex": q, "$options": "i"}

    tasks = list(tasks_col.find(task_query,
        sort=[("date", ASCENDING), ("created_at", ASCENDING)]))
    learnings = list(learnings_col.find(learning_query,
        sort=[("date", ASCENDING), ("created_at", ASCENDING)]))

    days = {}
    for t in tasks:
        d = t["date"]
        days.setdefault(d, {"tasks": [], "learnings": []})
        days[d]["tasks"].append(serialize_doc(t))
    for l in learnings:
        d = l["date"]
        days.setdefault(d, {"tasks": [], "learnings": []})
        days[d]["learnings"].append(serialize_doc(l))

    sorted_days = [{"date": d, **days[d]} for d in sorted(days.keys())]

    total_days_full = len(sorted_days)
    total_pages = max(1, (total_days_full + PAGE_SIZE - 1) // PAGE_SIZE)
    page = min(page, total_pages)
    start = (page - 1) * PAGE_SIZE
    paged_days = sorted_days[start: start + PAGE_SIZE]

    return jsonify({
        "days": paged_days,
        "total_tasks": len(tasks),
        "total_learnings": len(learnings),
        "total_days": total_days_full,
        "page": page,
        "total_pages": total_pages,
    })


# ─── Export ───────────────────────────────────────────────────────────────────

def _style_header_row(ws, headers, fill_hex):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color="FFFFFF", size=11)
    border_side = Side(style="thin", color="D0D5E0")
    border = Border(bottom=border_side)
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.border = border
    ws.row_dimensions[1].height = 22

def _style_data_row(ws, row_num, num_cols, even):
    fill = PatternFill("solid", fgColor="F7F8FC" if even else "FFFFFF")
    border_side = Side(style="thin", color="ECEEF3")
    border = Border(bottom=border_side)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

def _auto_col_widths(ws, min_w=12, max_w=60):
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(max_len + 2, min_w), max_w)

@app.route("/api/export")
@login_required
def export_excel():
    date_from = request.args.get("from")
    date_to   = request.args.get("to")
    if not date_from or not date_to:
        return jsonify({"error": "from and to dates required"}), 400
    if date_from > date_to:
        return jsonify({"error": "from must be <= to"}), 400

    tasks = list(tasks_col.find(
        {"date": {"$gte": date_from, "$lte": date_to}},
        sort=[("date", ASCENDING), ("created_at", ASCENDING)]
    ))
    learnings = list(learnings_col.find(
        {"date": {"$gte": date_from, "$lte": date_to}},
        sort=[("date", ASCENDING), ("created_at", ASCENDING)]
    ))

    wb = openpyxl.Workbook()

    # ── Tasks sheet ──
    ws_t = wb.active
    ws_t.title = "Tasks"
    ws_t.freeze_panes = "A2"
    t_headers = ["Date", "Title", "Note", "Status", "Created At"]
    _style_header_row(ws_t, t_headers, "6C63FF")
    for i, t in enumerate(tasks, 1):
        row = i + 1
        ws_t.cell(row=row, column=1, value=t.get("date", ""))
        ws_t.cell(row=row, column=2, value=t.get("title", ""))
        ws_t.cell(row=row, column=3, value=t.get("note", ""))
        ws_t.cell(row=row, column=4, value=t.get("status", "done"))
        ws_t.cell(row=row, column=5, value=t.get("created_at", "").isoformat() if isinstance(t.get("created_at"), datetime) else str(t.get("created_at", "")))
        _style_data_row(ws_t, row, len(t_headers), i % 2 == 0)
    _auto_col_widths(ws_t)

    # ── Learnings sheet ──
    ws_l = wb.create_sheet("Learnings")
    ws_l.freeze_panes = "A2"
    l_headers = ["Date", "Content", "Tags", "Created At"]
    _style_header_row(ws_l, l_headers, "0EA5E9")
    for i, l in enumerate(learnings, 1):
        row = i + 1
        ws_l.cell(row=row, column=1, value=l.get("date", ""))
        ws_l.cell(row=row, column=2, value=l.get("content", ""))
        ws_l.cell(row=row, column=3, value=", ".join(l.get("tags", [])))
        ws_l.cell(row=row, column=4, value=l.get("created_at", "").isoformat() if isinstance(l.get("created_at"), datetime) else str(l.get("created_at", "")))
        _style_data_row(ws_l, row, len(l_headers), i % 2 == 0)
    _auto_col_widths(ws_l)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"work-log_{date_from}_to_{date_to}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ─── Main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=os.environ.get("FLASK_ENV") != "production", use_reloader=True, reloader_type="stat")
